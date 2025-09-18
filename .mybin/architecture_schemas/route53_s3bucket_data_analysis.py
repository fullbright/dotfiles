#!/usr/bin/env python3
"""
Data Analysis Script for Route 53 and S3 Bucket Data
====================================================

This script processes CSV files containing Route 53 and S3 bucket data,
applies transformations, and generates Excel reports.

Author: Senior Python Developer
Version: 1.0
Requirements: openpyxl
"""

import os
import sys
import csv
import re
import logging
import argparse
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ColoredConsoleHandler(logging.StreamHandler):
    """Custom logging handler for colored console output."""
    
    COLORS = {
        'DEBUG': '\033[36m',    # Cyan
        'INFO': '\033[32m',     # Green
        'WARNING': '\033[33m',  # Yellow
        'ERROR': '\033[31m',    # Red
        'CRITICAL': '\033[35m', # Magenta
        'RESET': '\033[0m'      # Reset
    }

    def emit(self, record):
        color = self.COLORS.get(record.levelname, self.COLORS['RESET'])
        record.msg = f"{color}{record.msg}{self.COLORS['RESET']}"
        super().emit(record)


class DataAnalysisProcessor:
    """Main processor class for data analysis operations."""
    
    # Configuration constants
    ROUTE53_TYPES = {
        'rxdigitalplatform.co': 'rxdigitalplatform.com',
        'rxds-a_domains': 'rxds-a.com',
        'rxds-b_domains': 'rxds-b.com'
    }
    
    ROUTE53_FILTER_VALUES = ['NAME', 'CNAME']
    
    ROUTE53_COLUMNS = ["appli", "env", "module", "name", "ttl", "type", "value"]
    # "appli","env","module","name","ttl","type","value"
    S3_COLUMNS = ["account_name", "resourceid", "application", "environment", 
                  "versioningstatus", "encryption", "bucketpolicy"]
    # "account_name","resourceid","application","environment","versioningstatus","encryption","bucketpolicy"
    
    def __init__(self, config: Dict):
        """Initialize the processor with configuration."""
        self.config = config
        self.dry_run = config.get('dry_run', False)
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        self.logger = self._setup_logger()
        
        # Validate configuration
        self._validate_config()
    
    def _setup_logger(self) -> logging.Logger:
        """Set up logging configuration."""
        logger = logging.getLogger('DataAnalysisProcessor')
        logger.setLevel(logging.DEBUG)
        
        # Clear any existing handlers
        logger.handlers.clear()
        
        # Console handler with colors
        console_handler = ColoredConsoleHandler()
        console_handler.setLevel(logging.INFO)
        console_format = logging.Formatter(
            '%(asctime)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        console_handler.setFormatter(console_format)
        logger.addHandler(console_handler)
        
        # File handler
        log_file = Path(self.config['log_folder']) / f"data_analysis_{self.timestamp}.log"
        log_file.parent.mkdir(parents=True, exist_ok=True)
        
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)
        file_format = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        file_handler.setFormatter(file_format)
        logger.addHandler(file_handler)
        
        logger.info(f"Logger initialized. Log file: {log_file}")
        return logger
    
    def _validate_config(self):
        """Validate the configuration parameters."""
        required_keys = [
            's3_input_folder', 'route53_input_folder', 's3_output_folder',
            'route53_output_folder', 'log_folder'
        ]
        
        for key in required_keys:
            if key not in self.config:
                raise ValueError(f"Missing required configuration key: {key}")
            
            folder_path = Path(self.config[key])
            if not folder_path.exists():
                self.logger.warning(f"Creating directory: {folder_path}")
                folder_path.mkdir(parents=True, exist_ok=True)
    
    def find_latest_csv_files(self, folder_path: Path, expected_types: List[str]) -> Dict[str, Path]:
        """Find the latest CSV file for each type in the given folder."""
        self.logger.info(f"Searching for CSV files in: {folder_path}")
        
        files_by_type = {}
        csv_pattern = re.compile(r'^(.+)_(\d{8})\.csv$')
        
        if not folder_path.exists():
            self.logger.error(f"Folder does not exist: {folder_path}")
            return {}
        
        for file_path in folder_path.glob('*.csv'):
            match = csv_pattern.match(file_path.name)
            if match:
                file_type, date_str = match.groups()
                
                if file_type not in files_by_type or date_str > files_by_type[file_type][1]:
                    files_by_type[file_type] = (file_path, date_str)
        
        # Convert to final format and check for missing types
        result = {}
        for file_type in expected_types:
            if file_type in files_by_type:
                result[file_type] = files_by_type[file_type][0]
                self.logger.info(f"Found latest file for {file_type}: {result[file_type].name}")
            else:
                self.logger.warning(
                    f"🚨 WARNING: No CSV file found for type '{file_type}' in {folder_path}\n"
                    f"📝 To add/remove types from processing:\n"
                    f"   - For Route53: Modify ROUTE53_TYPES dictionary in the script\n"
                    f"   - For S3: Update expected_types list in process_s3_data method\n"
                    f"   - This type will be skipped and an empty sheet will be created if applicable"
                )
        
        return result
    
    def validate_csv_structure(self, file_path: Path, expected_columns: List[str]) -> bool:
        """Validate that CSV file has expected column structure."""
        try:
            # with open(file_path, 'r', newline='', encoding='utf-8') as f:
            with open(file_path, 'r', newline='', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                header = next(reader, None)
                
                if header is None:
                    self.logger.error(f"Empty CSV file: {file_path}")
                    return False
                
                # Check if all expected columns are present (case-insensitive)
                header_lower = [col.lower().strip() for col in header]
                expected_lower = [col.lower() for col in expected_columns]
                
                missing_columns = [col for col in expected_lower if col not in header_lower]
                if missing_columns:
                    self.logger.error(
                        f"CSV structure validation failed for {file_path}\n"
                        f"Current headers: {header_lower}\n"
                        f"Missing columns: {missing_columns}\n"
                        f"Expected columns: {expected_columns}\n"
                        f"Found columns: {header}"
                    )
                    return False
                
                self.logger.debug(f"CSV structure validated for {file_path}")
                return True
                
        except Exception as e:
            self.logger.error(f"Error validating CSV structure for {file_path}: {str(e)}")
            return False
    
    def read_csv_data(self, file_path: Path) -> List[Dict]:
        """Read CSV data and return as list of dictionaries."""
        data = []
        try:
            with open(file_path, 'r', newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # Clean up the row data
                    cleaned_row = {k.strip(): v.strip() if v else v for k, v in row.items()}
                    data.append(cleaned_row)
            
            self.logger.info(f"Read {len(data)} rows from {file_path}")
            return data
            
        except Exception as e:
            self.logger.error(f"Error reading CSV file {file_path}: {str(e)}")
            return []
    
    def filter_route53_data(self, data: List[Dict]) -> List[Dict]:
        """Filter Route53 data to keep only NAME and CNAME types."""
        filtered_data = [
            row for row in data 
            if row.get('type', '').upper() in self.ROUTE53_FILTER_VALUES
        ]
        
        self.logger.info(
            f"Filtered Route53 data: {len(data)} -> {len(filtered_data)} rows "
            f"(kept only {', '.join(self.ROUTE53_FILTER_VALUES)} types)"
        )
        return filtered_data
    
    def sort_s3_data(self, data: List[Dict]) -> List[Dict]:
        """Sort S3 data by account_name (asc), then environment (asc)."""
        def sort_key(row):
            account = row.get('account_name', '').lower()
            environment = row.get('environment', '').lower()
            return (account, environment)
        
        sorted_data = sorted(data, key=sort_key)
        self.logger.info(f"Sorted S3 data by account_name and environment ({len(sorted_data)} rows)")
        return sorted_data
    
    def apply_s3_transformations(self, data: List[Dict]) -> List[Dict]:
        """Apply transformations to S3 data. Placeholder for future enhancements."""
        # TODO: Add additional S3 data transformations here
        # Examples:
        # - Data cleansing
        # - Value standardization
        # - Additional filtering
        # - Calculated fields
        
        self.logger.debug("Applying S3 transformations (currently just sorting)")
        transformed_data = self.sort_s3_data(data)
        
        # Placeholder for future transformations
        # transformed_data = self._additional_s3_transform_1(transformed_data)
        # transformed_data = self._additional_s3_transform_2(transformed_data)
        
        return transformed_data
    
    def create_excel_workbook(self, sheet_names: List[str]) -> Workbook:
        """Create a new Excel workbook with specified sheet names."""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        for sheet_name in sheet_names:
            wb.create_sheet(title=sheet_name)
        
        return wb
    
    def write_data_to_sheet(self, wb: Workbook, sheet_name: str, data: List[Dict]):
        """Write data to a specific worksheet."""
        if sheet_name not in wb.sheetnames:
            self.logger.error(f"Sheet '{sheet_name}' not found in workbook")
            return
        
        ws = wb[sheet_name]
        
        if not data:
            self.logger.warning(f"No data to write to sheet '{sheet_name}'")
            return
        
        # Write headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        self.logger.info(f"Written {len(data)} rows to sheet '{sheet_name}'")
    
    def save_excel_file(self, wb: Workbook, file_path: Path):
        """Save Excel workbook to file."""
        try:
            if self.dry_run:
                dry_run_path = file_path.with_stem(f"{file_path.stem}_dryrun")
                wb.save(dry_run_path)
                self.logger.info(f"💾 DRY RUN: Excel file saved to {dry_run_path}")
            else:
                wb.save(file_path)
                self.logger.info(f"💾 Excel file saved to {file_path}")
        except Exception as e:
            self.logger.error(f"Error saving Excel file {file_path}: {str(e)}")
            raise
    
    def process_route53_data(self):
        """Process Route53 data and generate Excel reports."""
        self.logger.info("🚀 Starting Route53 data processing")
        
        input_folder = Path(self.config['route53_input_folder'])
        output_folder = Path(self.config['route53_output_folder'])
        
        # Find latest CSV files
        expected_types = list(self.ROUTE53_TYPES.keys())
        csv_files = self.find_latest_csv_files(input_folder, expected_types)
        
        # Create workbook
        sheet_names = list(self.ROUTE53_TYPES.values())
        wb = self.create_excel_workbook(sheet_names)
        
        # Process each type
        for csv_type, sheet_name in self.ROUTE53_TYPES.items():
            if csv_type in csv_files:
                csv_file = csv_files[csv_type]
                
                # Validate structure
                if not self.validate_csv_structure(csv_file, self.ROUTE53_COLUMNS):
                    self.logger.error(f"Stopping processing due to CSV structure validation failure")
                    sys.exit(1)
                
                # Read and process data
                data = self.read_csv_data(csv_file)
                filtered_data = self.filter_route53_data(data)
                
                # Write to sheet
                self.write_data_to_sheet(wb, sheet_name, filtered_data)
            else:
                self.logger.warning(f"🚨 Creating empty sheet '{sheet_name}' - no data file found")
        
        # Save files
        timestamped_file = output_folder / f"Route53_RxDS_{self.timestamp}.xlsx"
        regular_file = output_folder / "Route53_RxDS.xlsx"
        
        self.save_excel_file(wb, timestamped_file)
        self.save_excel_file(wb, regular_file)
        
        self.logger.info("✅ Route53 data processing completed")
    
    def process_s3_data(self):
        """Process S3 bucket data and generate Excel reports."""
        self.logger.info("🚀 Starting S3 data processing")
        
        input_folder = Path(self.config['s3_input_folder'])
        output_folder = Path(self.config['s3_output_folder'])
        
        # TODO: Define S3 expected types based on your requirements
        # For now, assuming all CSV files in the folder should be processed
        expected_types = []  # Add specific types if needed
        
        # Find all CSV files (modify this logic based on your S3 file naming convention)
        csv_files = {}
        csv_pattern = re.compile(r'^(.+)_(\d{8})\.csv$')
        
        if input_folder.exists():
            for file_path in input_folder.glob('*.csv'):
                match = csv_pattern.match(file_path.name)
                if match:
                    file_type, date_str = match.groups()
                    if file_type not in csv_files or date_str > csv_files[file_type][1]:
                        csv_files[file_type] = (file_path, date_str)
        
        if not csv_files:
            self.logger.warning("🚨 No S3 CSV files found")
            return
        
        # Process all S3 data into one combined dataset
        all_s3_data = []
        for file_type, (csv_file, _) in csv_files.items():
            self.logger.info(f"Processing S3 file: {csv_file.name}")
            
            # Validate structure
            if not self.validate_csv_structure(csv_file, self.S3_COLUMNS):
                self.logger.error(f"Stopping processing due to CSV structure validation failure")
                sys.exit(1)
            
            # Read data
            data = self.read_csv_data(csv_file)
            all_s3_data.extend(data)
        
        # Apply transformations
        transformed_data = self.apply_s3_transformations(all_s3_data)
        
        # Create workbook with single sheet
        wb = self.create_excel_workbook(['S3_Buckets'])
        self.write_data_to_sheet(wb, 'S3_Buckets', transformed_data)
        
        # Save files
        timestamped_file = output_folder / f"S3 Buckets  - Publishers & Consumers_{self.timestamp}.xlsx"
        regular_file = output_folder / "S3 Buckets  - Publishers & Consumers.xlsx"
        
        self.save_excel_file(wb, timestamped_file)
        self.save_excel_file(wb, regular_file)
        
        self.logger.info("✅ S3 data processing completed")
    
    def check_previous_excel_files(self):
        """Check for previous Excel files and log their status."""
        # Check previous Route53 file
        prev_route53 = self.config.get('prev_route53_excel')
        if prev_route53:
            prev_path = Path(prev_route53)
            if prev_path.exists():
                try:
                    # Try to open to check if corrupted
                    wb = load_workbook(prev_path)
                    wb.close()
                    self.logger.info(f"✅ Previous Route53 Excel file found and validated: {prev_path}")
                except Exception as e:
                    self.logger.error(f"❌ Previous Route53 Excel file is corrupted: {prev_path} - {str(e)}")
                    sys.exit(1)
            else:
                self.logger.warning(f"⚠️  Previous Route53 Excel file not found: {prev_path}")
        
        # Check previous S3 file
        prev_s3 = self.config.get('prev_s3_excel')
        if prev_s3:
            prev_path = Path(prev_s3)
            if prev_path.exists():
                try:
                    # Try to open to check if corrupted
                    wb = load_workbook(prev_path)
                    wb.close()
                    self.logger.info(f"✅ Previous S3 Excel file found and validated: {prev_path}")
                except Exception as e:
                    self.logger.error(f"❌ Previous S3 Excel file is corrupted: {prev_path} - {str(e)}")
                    sys.exit(1)
            else:
                self.logger.warning(f"⚠️  Previous S3 Excel file not found: {prev_path}")
    
    def run(self):
        """Main execution method."""
        self.logger.info("=" * 80)
        self.logger.info("🔄 STARTING DATA ANALYSIS SCRIPT")
        self.logger.info(f"🔧 Dry Run Mode: {'ENABLED' if self.dry_run else 'DISABLED'}")
        self.logger.info(f"⏰ Execution Timestamp: {self.timestamp}")
        self.logger.info("=" * 80)
        
        try:
            # Check previous Excel files
            self.check_previous_excel_files()
            
            # Process Route53 data
            self.process_route53_data()
            
            # Process S3 data
            self.process_s3_data()
            
            self.logger.info("=" * 80)
            self.logger.info("🎉 DATA ANALYSIS SCRIPT COMPLETED SUCCESSFULLY")
            self.logger.info("=" * 80)
            
        except Exception as e:
            self.logger.error(f"💥 SCRIPT FAILED: {str(e)}")
            sys.exit(1)


def get_default_config() -> Dict:
    """Get default configuration values."""
    script_dir = Path(__file__).parent.absolute()
    
    return {
        's3_input_folder': str(script_dir / 'input' / 's3_data'),
        'route53_input_folder': str(script_dir / 'input' / 'route53_data'),
        's3_output_folder': str(script_dir / 'output' / 's3_reports'),
        'route53_output_folder': str(script_dir / 'output' / 'route53_reports'),
        'log_folder': str(script_dir / 'logs'),
        'prev_route53_excel': '',
        'prev_s3_excel': '',
        'dry_run': False
    }


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description='Data Analysis Script for Route 53 and S3 Bucket Data',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python data_analysis.py --dry-run
  python data_analysis.py --s3-input /path/to/s3/data --route53-input /path/to/route53/data
  python data_analysis.py --prev-route53-excel /path/to/previous/route53.xlsx
        """
    )
    
    # Get default configuration
    defaults = get_default_config()
    
    # Add arguments
    parser.add_argument('--s3-input', default=defaults['s3_input_folder'],
                       help=f'S3 input folder (default: {defaults["s3_input_folder"]})')
    parser.add_argument('--route53-input', default=defaults['route53_input_folder'],
                       help=f'Route53 input folder (default: {defaults["route53_input_folder"]})')
    parser.add_argument('--s3-output', default=defaults['s3_output_folder'],
                       help=f'S3 output folder (default: {defaults["s3_output_folder"]})')
    parser.add_argument('--route53-output', default=defaults['route53_output_folder'],
                       help=f'Route53 output folder (default: {defaults["route53_output_folder"]})')
    parser.add_argument('--log-folder', default=defaults['log_folder'],
                       help=f'Log folder (default: {defaults["log_folder"]})')
    parser.add_argument('--prev-route53-excel', default=defaults['prev_route53_excel'],
                       help='Previous Route53 Excel file path')
    parser.add_argument('--prev-s3-excel', default=defaults['prev_s3_excel'],
                       help='Previous S3 Excel file path')
    parser.add_argument('--dry-run', action='store_true',
                       help='Perform dry run (append "_dryrun" to output files)')
    
    args = parser.parse_args()
    
    # Build configuration
    config = {
        's3_input_folder': args.s3_input,
        'route53_input_folder': args.route53_input,
        's3_output_folder': args.s3_output,
        'route53_output_folder': args.route53_output,
        'log_folder': args.log_folder,
        'prev_route53_excel': args.prev_route53_excel,
        'prev_s3_excel': args.prev_s3_excel,
        'dry_run': args.dry_run
    }
    
    # Run processor
    processor = DataAnalysisProcessor(config)
    processor.run()


if __name__ == '__main__':
    main()