import openpyxl
import json
import re
import logging
import shutil
from datetime import datetime
from pathlib import Path

# Configuration
EXCEL_FILE = r"C:\Users\AFANOUS\OneDrive - Luxottica Group S.p.A\Documenti\93.ftj.projects\302.096.PRJ.schemas_d_architecture\Kekeli_s_s3_bucket_changes_follow_up.xlsx"
SHEET_NAME = "raw_data"
TABLE_NAME = "tbl_s3_changes"

# Configurable lists (case-insensitive matching)
ENVIRONMENTS = ['dev', 'tin', 'cin', 'gin', 'qua', 'ppr', 'prd', 'apt']
PLATFORMS = ['gdp', 'spp', 'lsvw', 'eyec', 'ey', 'dbu']

# Setup logging
LOG_FILE = "logs/s3_parser.log"
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, mode='w', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


def clean_email_body(text):
    """Remove unwanted prefixes and suffixes from email body."""
    if not text:
        return text
    
    # Remove "WARNING: EXTERNAL EMAIL" prefix
    text = re.sub(r'^WARNING:\s*EXTERNAL\s+EMAIL\s*', '', text, flags=re.IGNORECASE | re.MULTILINE)
    
    # Remove everything after "-- If you wish to stop receiving"
    match = re.search(r'--\s*If you wish to stop receiv', text, flags=re.IGNORECASE)
    if match:
        text = text[:match.start()]
    
    return text.strip()


def flexible_json_parse(text):
    """Parse JSON with resilience to line breaks and formatting issues."""
    if not text:
        return None
    
    try:
        # First attempt: direct parse
        return json.loads(text)
    except json.JSONDecodeError:
        try:
            # Second attempt: remove excessive whitespace and line breaks
            cleaned = re.sub(r'\s+', ' ', text)
            return json.loads(cleaned)
        except json.JSONDecodeError:
            try:
                # Third attempt: fix common issues with line breaks in strings
                # Replace line breaks within quoted strings with spaces
                fixed = re.sub(r'(?<=")([^"]*?)[\r\n]+([^"]*?)(?=")', r'\1 \2', text)
                return json.loads(fixed)
            except json.JSONDecodeError as e:
                logger.error(f"JSON parsing failed: {e}")
                return None


def extract_keywords_from_bucket(bucket_name, keyword_list):
    """
    Extract keywords from bucket name using word boundary matching.
    Returns comma-separated string of matches or 'unknown'.
    """
    if not bucket_name:
        return 'unknown'
    
    # Split bucket name by dashes and other common separators
    parts = re.split(r'[-_.]', bucket_name.lower())
    
    matches = []
    for keyword in keyword_list:
        if keyword.lower() in parts:
            matches.append(keyword.lower())
    
    if matches:
        return ','.join(matches)
    return 'unknown'


def parse_iso_datetime(iso_string):
    """Convert ISO datetime string to Excel datetime format (timezone-naive)."""
    if not iso_string:
        return None
    
    try:
        # Parse ISO format: 2025-09-11T14:42:03Z
        # Excel doesn't support timezones, so we remove timezone info
        dt = datetime.fromisoformat(iso_string.replace('Z', '+00:00'))
        # Remove timezone info to make it naive
        dt_naive = dt.replace(tzinfo=None)
        return dt_naive
    except Exception as e:
        logger.warning(f"Failed to parse datetime '{iso_string}': {e}")
        return iso_string


def extract_data_from_json(json_data, bucket_name):
    """Extract required fields from JSON data."""
    extracted = {}
    
    try:
        detail = json_data.get('detail', {})
        
        # Extract environment and platform from bucket name
        extracted['environment'] = extract_keywords_from_bucket(bucket_name, ENVIRONMENTS)
        extracted['platform'] = extract_keywords_from_bucket(bucket_name, PLATFORMS)
        
        # Extract from JSON
        extracted['region'] = detail.get('awsRegion')
        extracted['bucket_name'] = bucket_name
        extracted['event_name'] = detail.get('eventName')
        extracted['operation_timestamp'] = parse_iso_datetime(detail.get('eventTime'))
        extracted['error_code'] = detail.get('errorCode')
        extracted['error_message'] = detail.get('errorMessage')
        extracted['aws_account'] = json_data.get('account')
        
    except Exception as e:
        logger.warning(f"Error extracting data: {e}")
    
    return extracted


def get_table_range(sheet, table_name):
    """Get the range of a table in the sheet."""
    for table in sheet.tables.values():
        if table.name == table_name:
            return table.ref
    return None


def process_excel_file():
    """Main function to process the Excel file."""
    logger.info("="*80)
    logger.info("Starting S3 Bucket Changes Parser")
    logger.info("="*80)
    
    # Check if file exists
    excel_path = Path(EXCEL_FILE)
    if not excel_path.exists():
        logger.error(f"Excel file not found: {EXCEL_FILE}")
        return False
    
    # Create backup before processing
    backup_path = excel_path.parent / f"{excel_path.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{excel_path.suffix}"
    try:
        logger.info(f"Creating backup: {backup_path.name}")
        shutil.copy2(EXCEL_FILE, backup_path)
        logger.info("Backup created successfully")
    except Exception as e:
        logger.warning(f"Could not create backup: {e}")
    
    try:
        # Load workbook
        logger.info(f"Loading workbook: {EXCEL_FILE}")
        wb = openpyxl.load_workbook(EXCEL_FILE)
        
        # Check if sheet exists
        if SHEET_NAME not in wb.sheetnames:
            logger.error(f"Sheet '{SHEET_NAME}' not found in workbook")
            return False
        
        sheet = wb[SHEET_NAME]
        logger.info(f"Found sheet: {SHEET_NAME}")
        
        # Get table range
        table_ref = get_table_range(sheet, TABLE_NAME)
        if not table_ref:
            logger.error(f"Table '{TABLE_NAME}' not found in sheet")
            return False
        
        logger.info(f"Found table: {TABLE_NAME} (range: {table_ref})")
        
        # Get header row
        min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(table_ref)
        
        # Read headers
        headers = {}
        for col_idx in range(min_col, max_col + 1):
            cell_value = sheet.cell(row=min_row, column=col_idx).value
            if cell_value:
                headers[cell_value] = col_idx
        
        logger.info(f"Headers found: {list(headers.keys())}")
        
        # Check required columns
        required_columns = ['email_body', 'environment', 'platform', 'region', 'bucket_name', 
                          'event_name', 'operation_timestamp', 'error_code', 'error_message', 'aws_account']
        
        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            logger.error(f"Missing required columns: {missing_columns}")
            return False
        
        # Process rows
        total_rows = max_row - min_row
        processed = 0
        skipped = 0
        errors = 0
        
        logger.info(f"Processing {total_rows} rows...")
        logger.info("-"*80)
        
        for row_idx in range(min_row + 1, max_row + 1):
            row_num = row_idx - min_row
            
            try:
                # Get email_body
                email_body_col = headers['email_body']
                email_body = sheet.cell(row=row_idx, column=email_body_col).value
                
                # Skip if email_body is empty
                if not email_body or str(email_body).strip() == '':
                    logger.warning(f"Row {row_num}/{total_rows}: Skipping - empty email_body")
                    skipped += 1
                    continue
                
                # Check if environment column already has data
                env_col = headers['environment']
                existing_env = sheet.cell(row=row_idx, column=env_col).value
                if existing_env and str(existing_env).strip() != '':
                    logger.info(f"Row {row_num}/{total_rows}: Skipping - already has environment data")
                    skipped += 1
                    continue
                
                # Clean email body
                cleaned_body = clean_email_body(email_body)
                
                # Parse JSON
                json_data = flexible_json_parse(cleaned_body)
                if not json_data:
                    logger.error(f"Row {row_num}/{total_rows}: Failed to parse JSON")
                    errors += 1
                    continue
                
                # Get bucket name from JSON
                bucket_name = json_data.get('detail', {}).get('requestParameters', {}).get('bucketName')
                if not bucket_name:
                    logger.warning(f"Row {row_num}/{total_rows}: No bucket name found in JSON")
                    errors += 1
                    continue
                
                # Extract data
                extracted_data = extract_data_from_json(json_data, bucket_name)
                
                # Write extracted data to cells
                for field_name, value in extracted_data.items():
                    if field_name in headers:
                        col_idx = headers[field_name]
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        
                        # Only write if cell is empty
                        if not cell.value or str(cell.value).strip() == '':
                            cell.value = value
                
                logger.info(f"Row {row_num}/{total_rows}: ✓ Processed - {bucket_name} | {extracted_data.get('environment')} | {extracted_data.get('platform')}")
                processed += 1
                
            except Exception as e:
                logger.error(f"Row {row_num}/{total_rows}: Unexpected error - {e}")
                errors += 1
                continue
        
        # Save workbook
        logger.info("-"*80)
        logger.info("Saving workbook...")
        try:
            wb.save(EXCEL_FILE)
            logger.info("Workbook saved successfully")
        except Exception as save_error:
            logger.error(f"Failed to save workbook: {save_error}")
            logger.info(f"Your original file is backed up at: {backup_path}")
            return False
        
        # Summary
        logger.info("="*80)
        logger.info("SUMMARY")
        logger.info("="*80)
        logger.info(f"Total rows:      {total_rows}")
        logger.info(f"Processed:       {processed}")
        logger.info(f"Skipped:         {skipped}")
        logger.info(f"Errors:          {errors}")
        logger.info("="*80)
        
        return True
        
    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        return False


if __name__ == "__main__":
    try:
        success = process_excel_file()
        if success:
            logger.info("Script completed successfully")
        else:
            logger.error("Script completed with errors")
            exit(1)
    except KeyboardInterrupt:
        logger.warning("Script interrupted by user")
        exit(1)
    except Exception as e:
        logger.error(f"Unexpected error: {e}", exc_info=True)
        exit(1)